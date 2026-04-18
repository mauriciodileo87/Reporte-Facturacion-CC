from __future__ import annotations

import os
import zipfile
from xml.etree import ElementTree as ET

try:
    from openpyxl import load_workbook as _openpyxl_load_workbook
    _TIENE_OPENPYXL = True
except ModuleNotFoundError:
    _openpyxl_load_workbook = None
    _TIENE_OPENPYXL = False


def tiene_openpyxl() -> bool:
    return _TIENE_OPENPYXL


def _col_ref_a_idx(cell_ref: str) -> int:
    letras = "".join(ch for ch in str(cell_ref or "") if ch.isalpha()).upper()
    idx = 0
    for ch in letras:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return max(idx - 1, 0)


def _parse_xlsx_shared_strings(zf: zipfile.ZipFile, ns: dict[str, str]) -> list[str]:
    try:
        root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    except KeyError:
        return []

    out: list[str] = []
    for si in root.findall("main:si", ns):
        textos = [node.text or "" for node in si.findall(".//main:t", ns)]
        out.append("".join(textos))
    return out


def _iter_internal_sheet_rows(path: str) -> list[list[str]]:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    with zipfile.ZipFile(path) as zf:
        workbook = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        shared_strings = _parse_xlsx_shared_strings(zf, ns)

        hoja = workbook.find("main:sheets/main:sheet", ns)
        if hoja is None:
            return []

        rel_id = hoja.attrib.get(f"{{{ns['rel']}}}id", "")
        target_by_id = {
            rel.attrib.get("Id", ""): rel.attrib.get("Target", "")
            for rel in rels.findall("pkgrel:Relationship", ns)
        }
        target = target_by_id.get(rel_id, "worksheets/sheet1.xml").lstrip("/")
        sheet_path = target if target.startswith("xl/") else f"xl/{target}"

        sheet_root = ET.fromstring(zf.read(sheet_path))
        rows: list[list[str]] = []
        for row_node in sheet_root.findall("main:sheetData/main:row", ns):
            fila: list[str] = []
            for cell in row_node.findall("main:c", ns):
                cell_ref = cell.attrib.get("r", "")
                idx = _col_ref_a_idx(cell_ref)
                while len(fila) <= idx:
                    fila.append("")

                cell_type = cell.attrib.get("t", "")
                if cell_type == "inlineStr":
                    valor = "".join(node.text or "" for node in cell.findall(".//main:t", ns))
                else:
                    v_node = cell.find("main:v", ns)
                    valor = v_node.text if v_node is not None and v_node.text is not None else ""
                    if cell_type == "s":
                        try:
                            valor = shared_strings[int(valor)]
                        except Exception:
                            valor = ""
                fila[idx] = valor
            rows.append(fila)
        return rows


def _read_with_openpyxl(path: str, read_only: bool) -> list[list[object]]:
    ext = os.path.splitext(path)[1].lower()
    wb = _openpyxl_load_workbook(path, data_only=True, read_only=read_only, keep_vba=(ext == ".xlsm"))
    ws = wb.worksheets[0]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def leer_excel_rows(path: str, *, read_only: bool = False) -> list[list[object]]:
    errores: list[str] = []

    if _TIENE_OPENPYXL:
        try:
            return _read_with_openpyxl(path, read_only=read_only)
        except Exception as exc:
            errores.append(f"openpyxl: {exc}")

    try:
        return _iter_internal_sheet_rows(path)
    except Exception as exc:
        errores.append(f"lector interno XLSX: {exc}")

    detalle = " | ".join(err for err in errores if err)
    raise ValueError(f"No se pudo leer el archivo Excel. {detalle}".strip())
